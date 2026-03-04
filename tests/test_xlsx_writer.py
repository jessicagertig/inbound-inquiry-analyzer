"""Tests for xlsx_writer.py: workbook structure, drop-downs, fills, and formatting."""
import pytest
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from inbound_inquiry_analyzer.config import CategoryConfig
from inbound_inquiry_analyzer.normalizer import NormalizedRecord
from inbound_inquiry_analyzer.xlsx_writer import generate_workbook

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

CATEGORIES = [
    {"name": "Issue with login", "color": "#FFD700"},
    {"name": "Inquiry about plan or pricing", "color": "#90EE90"},
    {"name": "Request for data deletion", "color": "#FFB6C1"},
    {"name": "Wrong email address (not intended for Polymer)", "color": "#ADD8E6"},
    {"name": "Advertising or solicitation", "color": "#D3D3D3"},
    {"name": "Knowledge base help", "color": "#DDA0DD"},
    {"name": "Feature help", "color": "#87CEEB"},
    {"name": "Migration", "color": "#F0E68C"},
    {"name": "Refund request", "color": "#FFA07A"},
    {"name": "Other", "color": "#E0E0E0"},
    {"name": "Unclear", "color": "#FF4500"},
]
SOURCES = ["Intercom", "In-app contact form", "Discord", "Direct email inquiry", "LinkedIn message"]


@pytest.fixture
def config():
    return CategoryConfig(categories=CATEGORIES, sources=SOURCES)


def make_record(
    inquiry_id="test-1",
    source="Intercom",
    date="2024-01-15",
    ts=1705276800.0,
    from_name="Jane Doe",
    from_email="jane@example.com",
    subject="Test subject",
    body="Test body",
) -> NormalizedRecord:
    return NormalizedRecord(
        inquiry_id=inquiry_id,
        source=source,
        received_at_date=date,
        received_at_ts=ts,
        from_name=from_name,
        from_email=from_email,
        subject=subject,
        message_body=body,
    )


@pytest.fixture
def sample_records():
    return [
        make_record("r1", ts=1705276800.0, subject="Login issue"),
        make_record("r2", ts=1705363200.0, subject="Pricing query"),
        make_record("r3", ts=1705190400.0, subject="Delete my data"),
    ]


@pytest.fixture
def sample_predictions():
    return [
        "Issue with login",
        "Inquiry about plan or pricing",
        "Unclear",
    ]


@pytest.fixture
def generated_wb(tmp_path, config, sample_records, sample_predictions):
    output = tmp_path / "test_output.xlsx"
    generate_workbook(sample_records, sample_predictions, config, output)
    return load_workbook(str(output))


# ---------------------------------------------------------------------------
# Structure tests
# ---------------------------------------------------------------------------

class TestWorkbookStructure:
    def test_inquiries_sheet_exists(self, generated_wb):
        assert "Inquiries" in generated_wb.sheetnames

    def test_lists_sheet_exists(self, generated_wb):
        assert "Lists" in generated_wb.sheetnames

    def test_eleven_columns(self, generated_wb):
        ws = generated_wb["Inquiries"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 12)]
        assert len([h for h in headers if h is not None]) == 11

    def test_column_headers_correct(self, generated_wb):
        ws = generated_wb["Inquiries"]
        expected = [
            "Inquiry ID",
            "Source",
            "Received At (YYYY-MM-DD)",
            "Received At (Timestamp)",
            "From Name",
            "From Email",
            "Subject",
            "Message Body",
            "Predicted Category",
            "Final Category",
            "Notes",
        ]
        for col_idx, expected_header in enumerate(expected, start=1):
            assert ws.cell(row=1, column=col_idx).value == expected_header

    def test_data_rows_count(self, generated_wb, sample_records):
        ws = generated_wb["Inquiries"]
        data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(v is not None for v in r)]
        assert len(data_rows) == len(sample_records)

    def test_rows_sorted_by_timestamp(self, generated_wb):
        ws = generated_wb["Inquiries"]
        # Column 4 is Received At (Timestamp)
        timestamps = [ws.cell(row=r, column=4).value for r in range(2, 5)]
        assert timestamps == sorted(timestamps)

    def test_final_category_matches_predicted_initially(self, generated_wb):
        ws = generated_wb["Inquiries"]
        for row in range(2, 5):
            predicted = ws.cell(row=row, column=9).value
            final = ws.cell(row=row, column=10).value
            assert predicted == final

    def test_notes_column_is_blank(self, generated_wb):
        ws = generated_wb["Inquiries"]
        for row in range(2, 5):
            assert ws.cell(row=row, column=11).value is None


# ---------------------------------------------------------------------------
# Formatting tests
# ---------------------------------------------------------------------------

class TestFormatting:
    def test_header_is_frozen(self, generated_wb):
        ws = generated_wb["Inquiries"]
        assert ws.freeze_panes is not None
        # Frozen at A2 means row 1 stays visible
        assert ws.freeze_panes == "A2"

    def test_auto_filter_enabled(self, generated_wb):
        ws = generated_wb["Inquiries"]
        assert ws.auto_filter.ref is not None

    def test_header_is_bold(self, generated_wb):
        ws = generated_wb["Inquiries"]
        for col in range(1, 12):
            cell = ws.cell(row=1, column=col)
            assert cell.font is not None and cell.font.bold, f"Column {col} header not bold"

    def test_header_has_fill(self, generated_wb):
        ws = generated_wb["Inquiries"]
        for col in range(1, 12):
            cell = ws.cell(row=1, column=col)
            assert cell.fill is not None and cell.fill.fill_type == "solid"

    def test_message_body_column_width_wider(self, generated_wb):
        ws = generated_wb["Inquiries"]
        msg_body_width = ws.column_dimensions["H"].width
        inquiry_id_width = ws.column_dimensions["A"].width
        assert msg_body_width > inquiry_id_width


# ---------------------------------------------------------------------------
# Data validation tests
# ---------------------------------------------------------------------------

class TestDataValidation:
    def test_final_category_has_data_validation(self, generated_wb):
        ws = generated_wb["Inquiries"]
        # Find any validation that covers column J (Final Category)
        dv_list = ws.data_validations.dataValidation
        found = False
        for dv in dv_list:
            if dv.formula1 and "Lists" in dv.formula1:
                for sqref in str(dv.sqref).split():
                    if "J" in sqref:
                        found = True
                        break
        assert found, "No data validation for Final Category (column J)"

    def test_source_has_data_validation(self, generated_wb):
        ws = generated_wb["Inquiries"]
        dv_list = ws.data_validations.dataValidation
        found = False
        for dv in dv_list:
            if dv.formula1 and "Lists" in dv.formula1:
                for sqref in str(dv.sqref).split():
                    if "B" in sqref:
                        found = True
                        break
        assert found, "No data validation for Source (column B)"


# ---------------------------------------------------------------------------
# Color fill tests
# ---------------------------------------------------------------------------

class TestColorFills:
    def test_predicted_category_cells_have_fills(self, generated_wb):
        ws = generated_wb["Inquiries"]
        for row in range(2, 5):
            cell = ws.cell(row=row, column=9)
            assert cell.fill is not None and cell.fill.fill_type == "solid"

    def test_final_category_cells_have_fills(self, generated_wb):
        ws = generated_wb["Inquiries"]
        for row in range(2, 5):
            cell = ws.cell(row=row, column=10)
            assert cell.fill is not None and cell.fill.fill_type == "solid"

    def test_unclear_category_has_distinct_fill(self, generated_wb, config):
        ws = generated_wb["Inquiries"]
        unclear_color = config.color_map["Unclear"].lstrip("#").upper()
        # Find the Unclear row (r3 has ts=1705190400.0 which is smallest -> row 2)
        # Check that at least one predicted category cell has the Unclear color
        found = False
        for row in range(2, 5):
            cell = ws.cell(row=row, column=9)
            if cell.value == "Unclear":
                fill_color = cell.fill.fgColor.rgb if cell.fill.fgColor else None
                # fgColor.rgb may have alpha prefix (FF + hex)
                if fill_color and unclear_color in fill_color.upper():
                    found = True
        assert found, "No Unclear row with distinct fill color found"

    def test_category_color_matches_config(self, generated_wb, config):
        ws = generated_wb["Inquiries"]
        for row in range(2, 5):
            category = ws.cell(row=row, column=9).value
            if category and category in config.color_map:
                expected_color = config.color_map[category].lstrip("#").upper()
                cell_fill = ws.cell(row=row, column=9).fill
                if cell_fill and cell_fill.fgColor:
                    actual_rgb = cell_fill.fgColor.rgb.upper()  # may have FF prefix
                    assert expected_color in actual_rgb, (
                        f"Row {row}: category {category!r} expected color {expected_color} "
                        f"but got {actual_rgb}"
                    )


# ---------------------------------------------------------------------------
# Non-configured category extension tests (T4)
# ---------------------------------------------------------------------------

class TestNonConfiguredCategories:
    """When Claude returns a category not in config, it should appear in the
    Lists sheet dropdown so the workbook is always self-consistent."""

    def test_non_configured_category_in_lists_sheet(self, tmp_path, config):
        records = [make_record("1")]
        predictions = ["BrandNewCategory"]  # not in config

        output = tmp_path / "out.xlsx"
        generate_workbook(records, predictions, config, output)
        wb = load_workbook(str(output))

        lists_ws = wb["Lists"]
        cat_values = [lists_ws.cell(row=i, column=1).value for i in range(1, 20)]
        cat_values = [v for v in cat_values if v is not None]
        assert "BrandNewCategory" in cat_values

    def test_non_configured_category_appended_after_configured(self, tmp_path, config):
        records = [make_record("1")]
        predictions = ["ZZZNewCat"]

        output = tmp_path / "out.xlsx"
        generate_workbook(records, predictions, config, output)
        wb = load_workbook(str(output))

        lists_ws = wb["Lists"]
        cat_values = [lists_ws.cell(row=i, column=1).value for i in range(1, 25)]
        cat_values = [v for v in cat_values if v is not None]

        # Configured categories come first
        configured = config.category_names
        for i, cat in enumerate(configured):
            assert cat_values[i] == cat, f"Position {i}: expected {cat!r}, got {cat_values[i]!r}"

        # Extra category is appended after
        assert cat_values[len(configured)] == "ZZZNewCat"

    def test_multiple_non_configured_categories_sorted(self, tmp_path, config):
        records = [make_record("1"), make_record("2"), make_record("3")]
        predictions = ["ZZZ", "AAA", "MMM"]

        output = tmp_path / "out.xlsx"
        generate_workbook(records, predictions, config, output)
        wb = load_workbook(str(output))

        lists_ws = wb["Lists"]
        configured_count = len(config.category_names)
        extra_values = [
            lists_ws.cell(row=configured_count + i, column=1).value
            for i in range(1, 4)
        ]
        extra_values = [v for v in extra_values if v is not None]
        assert extra_values == sorted(extra_values), "Extra categories must be in sorted order"
        assert set(extra_values) == {"AAA", "MMM", "ZZZ"}

    def test_non_configured_category_uses_default_fill(self, tmp_path, config):
        """A category not in config should get #E0E0E0 fill (gray default)."""
        records = [make_record("1")]
        predictions = ["UnknownCat"]

        output = tmp_path / "out.xlsx"
        generate_workbook(records, predictions, config, output)
        wb = load_workbook(str(output))

        ws = wb["Inquiries"]
        pred_cell = ws.cell(row=2, column=9)
        assert pred_cell.fill is not None and pred_cell.fill.fill_type == "solid"
        # Should be E0E0E0 (the default gray)
        fill_rgb = pred_cell.fill.fgColor.rgb.upper()
        assert "E0E0E0" in fill_rgb

    def test_configured_only_predictions_no_extras_in_lists(self, tmp_path, config):
        """If all predictions are configured, no extra rows added to Lists."""
        records = [make_record("1"), make_record("2")]
        predictions = ["Issue with login", "Refund request"]

        output = tmp_path / "out.xlsx"
        generate_workbook(records, predictions, config, output)
        wb = load_workbook(str(output))

        lists_ws = wb["Lists"]
        configured_count = len(config.category_names)
        # Row after last configured should be empty
        extra_val = lists_ws.cell(row=configured_count + 1, column=1).value
        assert extra_val is None

    def test_cat_range_covers_extra_categories(self, tmp_path, config):
        """The data validation formula must reference the extended range."""
        records = [make_record("1")]
        predictions = ["NewCatAlpha"]

        output = tmp_path / "out.xlsx"
        generate_workbook(records, predictions, config, output)
        wb = load_workbook(str(output))

        ws = wb["Inquiries"]
        dv_list = ws.data_validations.dataValidation
        # Find the category validation formula
        for dv in dv_list:
            if dv.formula1 and "Lists!$A$" in dv.formula1:
                # Extract the row number from the formula like "Lists!$A$1:$A$12"
                import re
                m = re.search(r"\$A\$(\d+)$", dv.formula1)
                if m:
                    row_count = int(m.group(1))
                    expected_count = len(config.category_names) + 1  # +1 for NewCatAlpha
                    assert row_count == expected_count, (
                        f"cat_range row count {row_count} != expected {expected_count}"
                    )
                    break
        else:
            pytest.fail("No category data validation formula found in workbook")
