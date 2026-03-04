"""Tests for CLI entry point: classification mode wiring, output messaging."""
import json
import sys
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

from inbound_inquiry_analyzer.cli import main

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

INTERCOM_PAYLOAD = json.dumps([
    {
        "id": "conv_001",
        "created_at": 1700000000,
        "subject": "I cannot log in",
        "first_message": "<p>I forgot my password</p>",
        "contact_email": "user@example.com",
        "contact_name": "Test User",
    }
])

# Suppress _load_dotenv so tests control env vars via monkeypatch without .env
# interference. Without this, a real .env in the project root will override
# monkeypatch.delenv() calls because dotenv runs inside main().
_NO_DOTENV = patch("inbound_inquiry_analyzer.cli._load_dotenv", lambda: None)


@pytest.fixture
def input_file(tmp_path):
    f = tmp_path / "input.json"
    f.write_text(INTERCOM_PAYLOAD)
    return f


@pytest.fixture
def output_path(tmp_path):
    return tmp_path / "out.xlsx"


# ---------------------------------------------------------------------------
# Keyword-only mode
# ---------------------------------------------------------------------------

class TestKeywordOnlyFlag:
    def test_keyword_only_exits_zero(self, input_file, output_path, monkeypatch):
        monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
        with _NO_DOTENV:
            rc = main([str(input_file), "-o", str(output_path), "--keyword-only"])
        assert rc == 0

    def test_keyword_only_reports_keyword_classifier(
        self, input_file, output_path, monkeypatch, capsys
    ):
        monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
        with _NO_DOTENV:
            main([str(input_file), "-o", str(output_path), "--keyword-only"])
        err = capsys.readouterr().err
        assert "keyword classifier" in err

    def test_keyword_only_does_not_call_anthropic(
        self, input_file, output_path, monkeypatch
    ):
        monkeypatch.setenv("ANTHROPIC_API_KEY", "sk-fake-key")
        with _NO_DOTENV, patch("inbound_inquiry_analyzer.claude_classifier.classify_with_claude") as mock_fn:
            main([str(input_file), "-o", str(output_path), "--keyword-only"])
            mock_fn.assert_not_called()


# ---------------------------------------------------------------------------
# Default mode (no API key → keyword fallback)
# ---------------------------------------------------------------------------

class TestDefaultModeNoKey:
    def test_no_key_exits_zero(self, input_file, output_path, monkeypatch):
        monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
        with _NO_DOTENV:
            rc = main([str(input_file), "-o", str(output_path)])
        assert rc == 0

    def test_no_key_reports_keyword_classifier(
        self, input_file, output_path, monkeypatch, capsys
    ):
        monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
        with _NO_DOTENV:
            main([str(input_file), "-o", str(output_path)])
        err = capsys.readouterr().err
        assert "keyword classifier" in err


# ---------------------------------------------------------------------------
# Claude mode (API key set, mock API call)
# Patch inbound_inquiry_analyzer.cli.get_client because cli.py imports it
# directly: `from inbound_inquiry_analyzer.api_client import get_client`.
# Patching the source module (api_client.get_client) would not intercept the
# already-imported name in the cli module.
# ---------------------------------------------------------------------------

class TestClaudeMode:
    def _make_mock_client(self, response_text: str):
        content_block = MagicMock()
        content_block.text = response_text
        msg = MagicMock()
        msg.content = [content_block]
        client = MagicMock()
        client.messages.create.return_value = msg
        return client

    def test_claude_mode_exits_zero(self, input_file, output_path, monkeypatch):
        mock_client = self._make_mock_client("Issue with login")
        with _NO_DOTENV, patch("inbound_inquiry_analyzer.cli.get_client", return_value=mock_client):
            rc = main([str(input_file), "-o", str(output_path)])
        assert rc == 0

    def test_claude_mode_reports_claude_api(
        self, input_file, output_path, monkeypatch, capsys
    ):
        mock_client = self._make_mock_client("Issue with login")
        with _NO_DOTENV, patch("inbound_inquiry_analyzer.cli.get_client", return_value=mock_client):
            main([str(input_file), "-o", str(output_path)])
        err = capsys.readouterr().err
        assert "Claude API" in err

    def test_new_category_from_claude_in_output_xlsx(
        self, input_file, output_path, monkeypatch
    ):
        """New categories suggested by Claude should appear in the XLSX Lists sheet."""
        from openpyxl import load_workbook

        mock_client = self._make_mock_client("Accessibility request")
        with _NO_DOTENV, patch("inbound_inquiry_analyzer.cli.get_client", return_value=mock_client):
            main([str(input_file), "-o", str(output_path)])

        wb = load_workbook(str(output_path))
        # New category must be findable somewhere in the workbook
        found = False
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                if any(cell == "Accessibility request" for cell in row if cell):
                    found = True
                    break
        assert found, "New Claude category not found in output XLSX"


# ---------------------------------------------------------------------------
# Error handling: no input
# ---------------------------------------------------------------------------

def test_no_input_exits_one(monkeypatch):
    monkeypatch.setattr("sys.stdin.isatty", lambda: True)
    with _NO_DOTENV:
        rc = main([])
    assert rc == 1
