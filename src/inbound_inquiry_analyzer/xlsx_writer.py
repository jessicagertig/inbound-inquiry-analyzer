"""XLSX workbook generation for classified inbound inquiries.

Produces an Excel workbook with:
- 'Inquiries' sheet: data rows with 11 columns, sorted by timestamp
- 'Lists' hidden sheet: category/source lists for drop-down validation
- Header formatting: bold, background fill, frozen pane, auto-filter
- Column widths tailored to content
- Data validation drop-downs on Final Category and Source columns
- Per-cell color fills on Predicted Category and Final Category columns
  (static fills — they will NOT auto-update if user changes Final Category in Excel)
"""
from __future__ import annotations

from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from inbound_inquiry_analyzer.config import CategoryConfig
from inbound_inquiry_analyzer.normalizer import NormalizedRecord

# Column definitions: (header_label, attribute_on_record_or_None, width)
# None means the column is computed at write time (Predicted Category, Final Category, Notes)
_COLUMNS = [
    ("Inquiry ID",               "inquiry_id",        15),
    ("Source",                   "source",            20),
    ("Received At (YYYY-MM-DD)", "received_at_date",  22),
    ("Received At (Timestamp)",  "received_at_ts",    22),
    ("From Name",                "from_name",         20),
    ("From Email",               "from_email",        28),
    ("Subject",                  "subject",           30),
    ("Message Body",             "message_body",      60),
    ("Predicted Category",       None,                25),
    ("Final Category",           None,                25),
    ("Notes",                    None,                20),
]

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="4472C4")  # Blue header
_HEADER_FONT = Font(bold=True, color="FFFFFF")

# Validation extends this many rows beyond the last data row so users can add more
_VALIDATION_EXTRA_ROWS = 100

# Column indices (1-based)
_COL_SOURCE = 2
_COL_PREDICTED = 9
_COL_FINAL = 10


def _hex_to_fill(hex_color: str) -> PatternFill:
    """Convert a hex color string (with or without #) to an openpyxl PatternFill."""
    color = hex_color.lstrip("#").upper()
    return PatternFill(fill_type="solid", fgColor=color)


def generate_workbook(
    records: Sequence[NormalizedRecord],
    predicted_categories: Sequence[str],
    config: CategoryConfig,
    output_path: str | Path = "inquiries_categorized.xlsx",
) -> Path:
    """Generate a formatted XLSX workbook from classified inquiry records.

    Args:
        records: Normalized inquiry records (unsorted; will be sorted by received_at_ts).
        predicted_categories: Predicted category string for each record (parallel sequence).
        config: Category configuration (names, colors, sources).
        output_path: Destination file path. Defaults to inquiries_categorized.xlsx.

    Returns:
        Path to the written XLSX file.
    """
    output_path = Path(output_path)

    # Sort records by timestamp ascending
    paired = sorted(zip(records, predicted_categories), key=lambda x: x[0].received_at_ts)

    wb = Workbook()
    ws = wb.active
    ws.title = "Inquiries"

    # --- Lists helper sheet (hidden) for data validation references ---
    lists_ws = wb.create_sheet("Lists")
    lists_ws.sheet_state = "hidden"

    category_names = config.category_names
    source_names = config.sources

    # Extend category list with any non-configured categories present in predictions
    # (e.g. Claude may return a valid but unconfigured category). These are appended
    # in sorted order so the dropdown always includes every value in the workbook.
    configured_set = set(category_names)
    extra_cats = sorted(
        {cat for cat in predicted_categories if cat not in configured_set}
    )
    extended_cats = category_names + extra_cats

    for i, cat in enumerate(extended_cats, start=1):
        lists_ws.cell(row=i, column=1, value=cat)
    for i, src in enumerate(source_names, start=1):
        lists_ws.cell(row=i, column=2, value=src)

    cat_range = f"Lists!$A$1:$A${len(extended_cats)}"
    src_range = f"Lists!$B$1:$B${len(source_names)}"

    # --- Header row ---
    for col_idx, (header, _, _) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter across all columns
    last_col_letter = get_column_letter(len(_COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col_letter}1"

    # --- Column widths ---
    for col_idx, (_, _, width) in enumerate(_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # Enable text wrap for Message Body column (column 8)
    msg_body_col = get_column_letter(8)

    # --- Data rows ---
    color_map = config.color_map
    unclear_fill = _hex_to_fill(color_map.get("Unclear", "#FF4500"))

    for row_offset, (rec, predicted) in enumerate(paired, start=2):
        row = row_offset

        # Populate columns
        ws.cell(row=row, column=1, value=rec.inquiry_id)
        ws.cell(row=row, column=2, value=rec.source)
        ws.cell(row=row, column=3, value=rec.received_at_date)
        ws.cell(row=row, column=4, value=rec.received_at_ts)
        ws.cell(row=row, column=5, value=rec.from_name)
        ws.cell(row=row, column=6, value=rec.from_email)
        ws.cell(row=row, column=7, value=rec.subject)

        msg_cell = ws.cell(row=row, column=8, value=rec.message_body)
        msg_cell.alignment = Alignment(wrap_text=True)

        ws.cell(row=row, column=9, value=predicted)   # Predicted Category
        ws.cell(row=row, column=10, value=predicted)  # Final Category (defaults to predicted)
        ws.cell(row=row, column=11, value=None)        # Notes (blank)

        # Color fills for Predicted Category (col 9)
        pred_fill = _hex_to_fill(color_map.get(predicted, "#E0E0E0"))
        ws.cell(row=row, column=9).fill = pred_fill

        # Color fill for Final Category (col 10) — same initial color
        ws.cell(row=row, column=10).fill = pred_fill

    # --- Data validation drop-downs ---
    num_data_rows = len(paired)
    last_data_row = 1 + num_data_rows  # header=1, data rows start at 2
    validation_end_row = last_data_row + _VALIDATION_EXTRA_ROWS

    if num_data_rows > 0 or True:  # Always add validation for future manual entries
        # Final Category drop-down (column 10 = J)
        final_cat_dv = DataValidation(
            type="list",
            formula1=cat_range,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid Category",
            error="Please select a category from the drop-down list.",
        )
        ws.add_data_validation(final_cat_dv)
        final_cat_dv.sqref = f"J2:J{validation_end_row}"

        # Source drop-down (column 2 = B)
        source_dv = DataValidation(
            type="list",
            formula1=src_range,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid Source",
            error="Please select a source from the drop-down list.",
        )
        ws.add_data_validation(source_dv)
        source_dv.sqref = f"B2:B{validation_end_row}"

    wb.save(output_path)
    return output_path
